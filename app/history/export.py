# ============================================================================
# FORD CAD — Call History Export
# ============================================================================
# Generate printable incident reports in HTML, PDF, CSV, XLSX.
# Uses the same inline-CSS pattern as app/reporting/renderer.py.
# ============================================================================

import csv
import io
import json
import logging
import traceback
from datetime import datetime
from html import escape as _h
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger("history.export")

ARTIFACT_DIR = Path("artifacts/history")


def _ensure_dir(incident_id: int) -> Path:
    d = ARTIFACT_DIR / str(incident_id)
    d.mkdir(parents=True, exist_ok=True)
    return d


def _ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _fmt_time(val: Optional[str]) -> str:
    """Format a timestamp for display — show HH:MM:SS if present."""
    if not val:
        return "—"
    # If full ISO, extract time portion
    if "T" in str(val):
        parts = str(val).split("T")
        return parts[1][:8] if len(parts) > 1 else val
    if " " in str(val):
        parts = str(val).split(" ")
        return parts[1][:8] if len(parts) > 1 else val
    return str(val)


# ============================================================================
# Base CSS (inline, self-contained for print / PDF / email)
# ============================================================================
_BASE_CSS = """
@page { size: letter; margin: 0.5in; }
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, 'Segoe UI', Roboto, Arial, sans-serif; font-size: 11px; color: #1a1a2e; line-height: 1.5; background: #fff; }
.page { max-width: 8.5in; margin: 0 auto; padding: 20px; }
.header { display: flex; justify-content: space-between; align-items: flex-start; border-bottom: 3px solid #c0392b; padding-bottom: 10px; margin-bottom: 16px; }
.header h1 { font-size: 16px; color: #c0392b; margin: 0; }
.header .meta { text-align: right; font-size: 10px; color: #666; }
.badge { display: inline-block; padding: 2px 8px; border-radius: 3px; font-size: 10px; font-weight: 600; text-transform: uppercase; }
.badge-open { background: #27ae60; color: #fff; }
.badge-closed { background: #7f8c8d; color: #fff; }
.badge-dispatched { background: #e67e22; color: #fff; }
.badge-issue { background: #e74c3c; color: #fff; }
.section { margin-bottom: 16px; }
.section-title { font-size: 12px; font-weight: 700; color: #2c3e50; border-bottom: 1px solid #ddd; padding-bottom: 4px; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 0.5px; }
.grid { display: grid; grid-template-columns: 1fr 1fr; gap: 6px 24px; }
.grid .label { font-weight: 600; color: #555; font-size: 10px; text-transform: uppercase; }
.grid .value { font-size: 11px; }
table { width: 100%; border-collapse: collapse; font-size: 10px; }
th { background: #2c3e50; color: #fff; padding: 5px 8px; text-align: left; font-weight: 600; font-size: 10px; }
td { padding: 4px 8px; border-bottom: 1px solid #eee; }
tr:nth-child(even) { background: #f8f9fa; }
.narrative-entry { padding: 6px 0; border-bottom: 1px solid #eee; }
.narrative-entry .ts { font-size: 10px; color: #888; font-family: monospace; }
.narrative-entry .author { font-weight: 600; color: #2c3e50; }
.narrative-entry .text { margin-top: 2px; }
.issue-banner { background: #fdeaea; border: 1px solid #e74c3c; border-radius: 4px; padding: 8px 12px; color: #c0392b; font-weight: 600; margin-bottom: 12px; }
.footer { border-top: 1px solid #ddd; padding-top: 8px; margin-top: 20px; font-size: 9px; color: #999; text-align: center; }
.unit-timeline { margin: 4px 0; }
.unit-chip { display: inline-block; padding: 1px 6px; border-radius: 3px; font-size: 10px; font-weight: 600; background: #34495e; color: #fff; margin-right: 4px; }
@media print { body { font-size: 10px; } .page { padding: 0; } }
"""


# ============================================================================
# HTML Incident Report
# ============================================================================

def _is_http_audit(event_type: str) -> bool:
    """Return True for raw HTTP audit trail entries that clutter reports."""
    if not event_type:
        return False
    et = event_type.strip().upper()
    return et.startswith("HTTP_")


def render_incident_html(incident: Dict[str, Any]) -> str:
    """Render a self-contained, printable HTML incident report."""
    inc = incident
    narrative_entries = inc.get("narrative_entries", [])
    units = inc.get("units", [])
    history = [h for h in inc.get("history", []) if not _is_http_audit(h.get("event_type", ""))]

    run_num = _h(str(inc.get("incident_number") or inc.get("run_number") or inc.get("incident_id", "")))
    inc_type = _h(str(inc.get("type") or "Unknown"))
    location = _h(str(inc.get("location") or inc.get("address") or ""))
    status = str(inc.get("status") or "").upper()
    status_class = "badge-closed" if status == "CLOSED" else "badge-open" if status == "OPEN" else "badge-dispatched"

    issue = inc.get("issue_found") or inc.get("issue_flag") or 0
    issue_banner = ""
    if issue:
        issue_banner = '<div class="issue-banner">ISSUE FOUND — This incident has been flagged for review.</div>'

    # Header
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Incident Report — {run_num} — Ford Fire Department</title>
<style>{_BASE_CSS}
.print-toolbar {{
    position: sticky; top: 0; z-index: 100;
    display: flex; align-items: center; gap: 10px;
    background: #1e293b; color: #f8fafc;
    padding: 10px 20px; margin: 0;
    border-bottom: 2px solid #3b82f6;
    font-family: -apple-system, 'Segoe UI', Roboto, Arial, sans-serif;
}}
.print-toolbar button {{
    padding: 6px 16px; border-radius: 4px; font-size: 12px;
    font-weight: 600; cursor: pointer; border: 1px solid #64748b;
}}
.print-toolbar .btn-print {{
    background: #3b82f6; color: #fff; border-color: #3b82f6;
}}
.print-toolbar .btn-print:hover {{ opacity: 0.9; }}
.print-toolbar .btn-close {{
    background: #334155; color: #e2e8f0; border-color: #64748b;
}}
.print-toolbar .btn-close:hover {{ background: #475569; }}
.print-toolbar span {{ font-size: 13px; font-weight: 600; }}
@media print {{ .print-toolbar {{ display: none !important; }} }}
</style>
</head>
<body>
<div class="print-toolbar">
    <button class="btn-print" onclick="window.print()">Print</button>
    <button class="btn-close" onclick="window.close()">Close</button>
    <span>Incident Report #{run_num} — {inc_type}</span>
</div>
<div class="page">
<div class="header">
    <div>
        <h1>INCIDENT REPORT</h1>
        <div style="font-size:13px;font-weight:700;margin-top:4px;">#{run_num} — {inc_type}</div>
    </div>
    <div class="meta">
        Ford Fire Department CAD<br>
        Generated: {_ts()}<br>
        <span class="badge {status_class}">{_h(status)}</span>
    </div>
</div>

{issue_banner}

<div class="section">
    <div class="section-title">Summary</div>
    <div class="grid">
        <div><span class="label">Run #</span><br><span class="value">{run_num}</span></div>
        <div><span class="label">Type / Nature</span><br><span class="value">{inc_type}</span></div>
        <div><span class="label">Location</span><br><span class="value">{location}</span></div>
        <div><span class="label">Priority</span><br><span class="value">{_h(str(inc.get('priority') or '—'))}</span></div>
        <div><span class="label">Caller</span><br><span class="value">{_h(str(inc.get('caller_name') or '—'))}</span></div>
        <div><span class="label">Caller Phone</span><br><span class="value">{_h(str(inc.get('caller_phone') or '—'))}</span></div>
        <div><span class="label">Shift</span><br><span class="value">{_h(str(inc.get('shift') or '—'))}</span></div>
        <div><span class="label">Disposition</span><br><span class="value">{_h(str(inc.get('final_disposition') or '—'))}</span></div>
    </div>
</div>

<div class="section">
    <div class="section-title">Times</div>
    <div class="grid">
        <div><span class="label">Created</span><br><span class="value">{_h(str(inc.get('created') or '—'))}</span></div>
        <div><span class="label">Closed</span><br><span class="value">{_h(str(inc.get('closed_at') or '—'))}</span></div>
        <div><span class="label">Hold Start</span><br><span class="value">{_h(str(inc.get('held_at') or '—'))}</span></div>
        <div><span class="label">Hold End</span><br><span class="value">{_h(str(inc.get('held_released_at') or '—'))}</span></div>
    </div>
</div>
"""

    # Narrative
    if narrative_entries:
        html += '<div class="section"><div class="section-title">Narrative</div>\n'
        for entry in narrative_entries:
            ts = _h(str(entry.get("timestamp") or ""))
            author = _h(str(entry.get("user") or ""))
            text = _h(str(entry.get("text") or ""))
            html += f"""<div class="narrative-entry">
    <span class="ts">{ts}</span> <span class="author">{author}</span>
    <div class="text">{text}</div>
</div>\n"""
        html += "</div>\n"
    elif inc.get("narrative"):
        html += f"""<div class="section">
    <div class="section-title">Narrative</div>
    <div class="narrative-entry"><div class="text">{_h(str(inc['narrative']))}</div></div>
</div>\n"""

    # Units
    if units:
        html += """<div class="section">
    <div class="section-title">Units</div>
    <table>
        <tr><th>Unit</th><th>Assigned</th><th>Dispatched</th><th>Enroute</th><th>Arrived</th><th>Transporting</th><th>At Medical</th><th>Cleared</th><th>Disposition</th></tr>
"""
        for u in units:
            uid = _h(str(u.get("unit_id") or ""))
            html += f"""        <tr>
            <td><span class="unit-chip">{uid}</span></td>
            <td>{_fmt_time(u.get('assigned'))}</td>
            <td>{_fmt_time(u.get('dispatched'))}</td>
            <td>{_fmt_time(u.get('enroute'))}</td>
            <td>{_fmt_time(u.get('arrived'))}</td>
            <td>{_fmt_time(u.get('transporting'))}</td>
            <td>{_fmt_time(u.get('at_medical'))}</td>
            <td>{_fmt_time(u.get('cleared'))}</td>
            <td>{_h(str(u.get('disposition') or '—'))}</td>
        </tr>\n"""
        html += "    </table>\n</div>\n"

    # History / Timeline
    if history:
        html += """<div class="section">
    <div class="section-title">Timeline / Audit</div>
    <table>
        <tr><th>Time</th><th>Event</th><th>User</th><th>Unit</th><th>Details</th></tr>
"""
        for h in history:
            html += f"""        <tr>
            <td style="white-space:nowrap">{_h(str(h.get('timestamp') or ''))}</td>
            <td>{_h(str(h.get('event_type') or ''))}</td>
            <td>{_h(str(h.get('user') or ''))}</td>
            <td>{_h(str(h.get('unit_id') or ''))}</td>
            <td style="max-width:300px;overflow:hidden;text-overflow:ellipsis">{_h(str(h.get('details') or '')[:200])}</td>
        </tr>\n"""
        html += "    </table>\n</div>\n"

    # Footer
    html += f"""<div class="footer">
    Ford Fire Department — Computer-Aided Dispatch — Incident Report #{run_num} — Generated {_ts()}
</div>
</div>
</body>
</html>"""

    return html


# ============================================================================
# PDF via WeasyPrint
# ============================================================================

def render_incident_pdf(incident: Dict[str, Any], output_path: Path) -> bool:
    """Render an incident to PDF using WeasyPrint. Returns True on success."""
    try:
        from weasyprint import HTML as WeasyprintHTML
        html_string = render_incident_html(incident)
        WeasyprintHTML(string=html_string).write_pdf(str(output_path))
        logger.info("PDF saved: %s", output_path)
        return True
    except ImportError:
        logger.warning("WeasyPrint not installed — PDF export unavailable")
        return False
    except Exception:
        logger.error("PDF render failed:\n%s", traceback.format_exc())
        return False


# ============================================================================
# CSV
# ============================================================================

def render_incident_csv(incident: Dict[str, Any], output_path: Path) -> bool:
    """Render incident data as a flat CSV."""
    try:
        inc = incident
        units = inc.get("units", [])

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Summary section
            writer.writerow(["Incident Report"])
            writer.writerow([])
            summary_fields = [
                "incident_id", "incident_number", "run_number", "type", "location", "address",
                "status", "priority", "caller_name", "caller_phone", "shift",
                "final_disposition", "created", "closed_at", "issue_found",
            ]
            writer.writerow(["Field", "Value"])
            for field in summary_fields:
                writer.writerow([field, inc.get(field, "")])

            # Units section
            writer.writerow([])
            writer.writerow(["Units"])
            unit_headers = ["unit_id", "assigned", "dispatched", "enroute", "arrived",
                            "transporting", "at_medical", "cleared", "disposition"]
            writer.writerow(unit_headers)
            for u in units:
                writer.writerow([u.get(h, "") for h in unit_headers])

            # Narrative
            writer.writerow([])
            writer.writerow(["Narrative"])
            writer.writerow(["Timestamp", "User", "Text"])
            for n in inc.get("narrative_entries", []):
                writer.writerow([n.get("timestamp", ""), n.get("user", ""), n.get("text", "")])

        logger.info("CSV saved: %s", output_path)
        return True
    except Exception:
        logger.error("CSV render failed:\n%s", traceback.format_exc())
        return False


# ============================================================================
# XLSX
# ============================================================================

def render_incident_xlsx(incident: Dict[str, Any], output_path: Path) -> bool:
    """Render incident as multi-sheet Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        inc = incident

        # --- Summary sheet ---
        ws = wb.active
        ws.title = "Summary"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        ws.append(["Field", "Value"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        summary_fields = [
            ("Incident ID", inc.get("incident_id", "")),
            ("Run Number", inc.get("incident_number") or inc.get("run_number", "")),
            ("Type", inc.get("type", "")),
            ("Location", inc.get("location") or inc.get("address", "")),
            ("Status", inc.get("status", "")),
            ("Priority", inc.get("priority", "")),
            ("Caller", inc.get("caller_name", "")),
            ("Phone", inc.get("caller_phone", "")),
            ("Shift", inc.get("shift", "")),
            ("Disposition", inc.get("final_disposition", "")),
            ("Created", inc.get("created", "")),
            ("Closed", inc.get("closed_at", "")),
            ("Issue Found", "Yes" if inc.get("issue_found") else "No"),
        ]
        for label, val in summary_fields:
            ws.append([label, val])

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40

        # --- Units sheet ---
        ws2 = wb.create_sheet("Units")
        unit_headers = ["Unit", "Assigned", "Dispatched", "Enroute", "Arrived",
                        "Transporting", "At Medical", "Cleared", "Disposition"]
        ws2.append(unit_headers)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        for u in inc.get("units", []):
            ws2.append([
                u.get("unit_id", ""), u.get("assigned", ""), u.get("dispatched", ""),
                u.get("enroute", ""), u.get("arrived", ""), u.get("transporting", ""),
                u.get("at_medical", ""), u.get("cleared", ""), u.get("disposition", ""),
            ])

        # --- Narrative sheet ---
        ws3 = wb.create_sheet("Narrative")
        ws3.append(["Timestamp", "User", "Text"])
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
        for n in inc.get("narrative_entries", []):
            ws3.append([n.get("timestamp", ""), n.get("user", ""), n.get("text", "")])
        ws3.column_dimensions["C"].width = 60

        # --- Timeline sheet ---
        ws4 = wb.create_sheet("Timeline")
        ws4.append(["Timestamp", "Event", "User", "Unit", "Details"])
        for cell in ws4[1]:
            cell.font = header_font
            cell.fill = header_fill
        for h in inc.get("history", []):
            if _is_http_audit(h.get("event_type", "")):
                continue
            ws4.append([
                h.get("timestamp", ""), h.get("event_type", ""),
                h.get("user", ""), h.get("unit_id", ""), str(h.get("details", ""))[:200],
            ])
        ws4.column_dimensions["E"].width = 50

        wb.save(str(output_path))
        logger.info("XLSX saved: %s", output_path)
        return True
    except ImportError:
        logger.warning("openpyxl not installed — XLSX export unavailable")
        return False
    except Exception:
        logger.error("XLSX render failed:\n%s", traceback.format_exc())
        return False


# ============================================================================
# Convenience: render all requested formats
# ============================================================================

def render_incident_report(
    incident: Dict[str, Any],
    formats: List[str],
) -> Dict[str, str]:
    """Render an incident report in the requested formats.

    Returns dict mapping format -> absolute file path of the artifact.
    """
    incident_id = incident.get("incident_id", 0)
    artifact_dir = _ensure_dir(incident_id)
    results: Dict[str, str] = {}
    ts_slug = datetime.now().strftime("%Y%m%d_%H%M%S")

    if "html" in formats or "pdf" in formats:
        html_string = render_incident_html(incident)

        if "html" in formats:
            out = artifact_dir / f"incident_{incident_id}_{ts_slug}.html"
            out.write_text(html_string, encoding="utf-8")
            results["html"] = str(out.resolve())

        if "pdf" in formats:
            out = artifact_dir / f"incident_{incident_id}_{ts_slug}.pdf"
            if render_incident_pdf(incident, out):
                results["pdf"] = str(out.resolve())

    if "csv" in formats:
        out = artifact_dir / f"incident_{incident_id}_{ts_slug}.csv"
        if render_incident_csv(incident, out):
            results["csv"] = str(out.resolve())

    if "xlsx" in formats:
        out = artifact_dir / f"incident_{incident_id}_{ts_slug}.xlsx"
        if render_incident_xlsx(incident, out):
            results["xlsx"] = str(out.resolve())

    return results
