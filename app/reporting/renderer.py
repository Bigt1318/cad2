# ============================================================================
# FORD CAD - Multi-Format Report Renderer
# ============================================================================
# Renders report data (Python dicts with rows, stats, metadata) into multiple
# output formats: PDF, CSV, XLSX, HTML, TXT.
#
# Artifacts are saved to artifacts/reports/{run_id}/ via ensure_artifact_dir().
# ============================================================================

import base64
import csv
import io
import logging
import traceback
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
from html import escape as html_escape
from zoneinfo import ZoneInfo

from .models import ensure_artifact_dir, ARTIFACT_DIR

logger = logging.getLogger("reporting.renderer")

EASTERN = ZoneInfo("America/New_York")

# ---------------------------------------------------------------------------
# Load Ford logo as base64 for embedding in reports
# ---------------------------------------------------------------------------
_LOGO_B64 = ""
try:
    _logo_path = Path(__file__).resolve().parent.parent.parent / "static" / "images" / "logo.png"
    if _logo_path.exists():
        _LOGO_B64 = base64.b64encode(_logo_path.read_bytes()).decode("ascii")
        logger.info("Ford logo loaded from %s (%d bytes)", _logo_path, len(_LOGO_B64))
except Exception:
    logger.debug("Could not load Ford logo", exc_info=True)

# ---------------------------------------------------------------------------
# Import chart generator (optional)
# ---------------------------------------------------------------------------
try:
    from . import charts as _charts
    HAS_CHARTS = _charts.HAS_MATPLOTLIB
except ImportError:
    _charts = None  # type: ignore[assignment]
    HAS_CHARTS = False

# ---------------------------------------------------------------------------
# Shared CSS & HTML fragments  (inline for self-contained output)
# ---------------------------------------------------------------------------

_FORD_BLUE = "#003478"
_BRAND_BLUE_DARK = "#003478"
_BRAND_BLUE_LIGHT = "#1e5cb3"
_BRAND_GREEN = "#16a34a"
_BRAND_RED = "#dc2626"
_BRAND_YELLOW = "#ca8a04"
_GRAY_50 = "#f9fafb"
_GRAY_100 = "#f3f4f6"
_GRAY_200 = "#e5e7eb"
_GRAY_500 = "#6b7280"
_GRAY_700 = "#374151"

_BASE_CSS = f"""
    * {{ box-sizing: border-box; }}
    body {{
        font-family: Arial, Helvetica, sans-serif;
        margin: 0;
        padding: 0;
        background: {_GRAY_50};
        color: #111827;
        font-size: 13px;
        line-height: 1.5;
    }}
    .page-wrap {{
        max-width: 960px;
        margin: 0 auto;
        padding: 20px;
    }}
    .card {{
        background: #fff;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        overflow: hidden;
        margin-bottom: 24px;
    }}
    .header {{
        background: linear-gradient(135deg, {_FORD_BLUE}, {_BRAND_BLUE_LIGHT});
        color: #fff;
        padding: 24px 28px;
    }}
    .header h1 {{ margin: 0; font-size: 22px; }}
    .header .subtitle {{ margin: 4px 0 0; opacity: 0.9; font-size: 15px; }}
    .header-logo {{ height: 40px; margin-right: 16px; vertical-align: middle; }}
    .header-top {{ display: flex; align-items: center; margin-bottom: 6px; }}
    .header-org {{ font-size: 11px; opacity: 0.85; text-transform: uppercase; letter-spacing: 0.5px; }}
    .meta-bar {{
        display: flex; flex-wrap: wrap; gap: 16px;
        padding: 12px 28px; background: {_GRAY_100};
        font-size: 11px; color: {_GRAY_700};
        border-bottom: 1px solid {_GRAY_200};
    }}
    .meta-bar strong {{ color: {_FORD_BLUE}; }}
    .kpi-row {{
        display: flex; flex-wrap: wrap; gap: 14px; margin-bottom: 24px;
    }}
    .kpi-card {{
        flex: 1; min-width: 130px; max-width: 200px;
        border: 1px solid {_GRAY_200}; border-radius: 10px;
        padding: 16px; text-align: center; background: #fff;
    }}
    .kpi-card .kpi-value {{
        font-size: 28px; font-weight: 800; color: {_FORD_BLUE};
    }}
    .kpi-card .kpi-label {{
        font-size: 10px; color: {_GRAY_500}; text-transform: uppercase;
        letter-spacing: 0.3px; margin-top: 4px;
    }}
    .kpi-card .kpi-trend {{
        font-size: 11px; font-weight: 700; margin-top: 4px;
    }}
    .kpi-trend.up {{ color: {_BRAND_RED}; }}
    .kpi-trend.down {{ color: {_BRAND_GREEN}; }}
    .kpi-trend.flat {{ color: {_GRAY_500}; }}
    .chart-section {{
        margin-bottom: 28px; page-break-inside: avoid;
    }}
    .chart-section h3 {{
        color: {_FORD_BLUE}; font-size: 15px; margin-bottom: 8px;
        border-bottom: 2px solid {_GRAY_200}; padding-bottom: 6px;
    }}
    .chart-section .chart-narrative {{
        font-size: 12px; color: {_GRAY_500}; margin-top: 6px; font-style: italic;
    }}
    .chart-section img {{
        max-width: 100%; height: auto; display: block; margin: 0 auto;
    }}
    .chart-grid {{
        display: grid; grid-template-columns: repeat(2, 1fr);
        gap: 20px; margin-bottom: 24px;
    }}
    .branded-footer {{
        text-align: center; padding: 14px 28px;
        border-top: 2px solid {_FORD_BLUE}; color: {_GRAY_500};
        font-size: 10px;
    }}
    .branded-footer .confidential {{
        font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px;
        color: {_FORD_BLUE};
    }}
    .section-divider {{
        border: none; border-top: 3px solid {_FORD_BLUE};
        margin: 32px 0;
    }}
    .recommendation-box {{
        border-left: 4px solid {_BRAND_YELLOW};
        background: #fffbeb; padding: 12px 16px;
        border-radius: 0 8px 8px 0; margin-bottom: 8px;
        font-size: 12px;
    }}
    .body {{ padding: 24px 28px; }}
    .filter-bar {{
        display: flex;
        flex-wrap: wrap;
        gap: 12px;
        margin-bottom: 20px;
    }}
    .filter-chip {{
        background: {_GRAY_100};
        padding: 6px 14px;
        border-radius: 6px;
        font-size: 12px;
    }}
    .filter-chip strong {{ color: {_BRAND_BLUE_DARK}; }}
    table {{
        width: 100%;
        border-collapse: collapse;
        margin-bottom: 16px;
        font-size: 12px;
    }}
    th {{
        background: {_BRAND_BLUE_DARK};
        color: #fff;
        padding: 8px 10px;
        text-align: left;
        font-weight: 600;
        font-size: 11px;
        text-transform: uppercase;
        letter-spacing: 0.3px;
    }}
    td {{
        padding: 7px 10px;
        border-bottom: 1px solid {_GRAY_200};
    }}
    tr:nth-child(even) {{ background: #f8fafc; }}
    tr:hover {{ background: #eff6ff; }}
    .stat-grid {{
        display: grid;
        grid-template-columns: repeat(auto-fill, minmax(160px, 1fr));
        gap: 14px;
        margin-bottom: 24px;
    }}
    .stat-box {{
        border: 1px solid {_GRAY_200};
        border-radius: 8px;
        padding: 14px;
        text-align: center;
    }}
    .stat-box .value {{
        font-size: 26px;
        font-weight: 700;
        color: {_BRAND_BLUE_DARK};
    }}
    .stat-box .label {{
        font-size: 11px;
        color: {_GRAY_500};
        margin-top: 2px;
    }}
    .section-title {{
        color: {_GRAY_700};
        border-bottom: 2px solid {_GRAY_200};
        padding-bottom: 8px;
        margin: 28px 0 14px;
        font-size: 16px;
    }}
    .incident-card {{
        border: 1px solid {_GRAY_200};
        border-radius: 8px;
        padding: 14px;
        margin-bottom: 12px;
    }}
    .incident-card.issue {{ border-left: 4px solid {_BRAND_RED}; background: #fef2f2; }}
    .badge {{
        display: inline-block;
        padding: 2px 8px;
        border-radius: 4px;
        font-size: 10px;
        font-weight: 600;
    }}
    .badge-red {{ background: {_BRAND_RED}; color: #fff; }}
    .badge-green {{ background: #dcfce7; color: {_BRAND_GREEN}; }}
    .badge-yellow {{ background: #fef9c3; color: {_BRAND_YELLOW}; }}
    .badge-gray {{ background: {_GRAY_200}; color: {_GRAY_700}; }}
    .pass {{ color: {_BRAND_GREEN}; font-weight: 700; }}
    .fail {{ color: {_BRAND_RED}; font-weight: 700; }}
    .footer {{
        text-align: center;
        color: {_GRAY_500};
        font-size: 11px;
        padding: 16px 28px;
        border-top: 1px solid {_GRAY_200};
    }}
    @media print {{
        body {{ background: #fff; }}
        .page-wrap {{ padding: 0; max-width: 100%; }}
        .card {{ box-shadow: none; border: 1px solid #ccc; }}
    }}
    @page {{
        size: letter;
        margin: 0.6in 0.5in;
        @bottom-left {{
            content: "CONFIDENTIAL - Ford Motor Company";
            font-size: 8px;
            color: {_FORD_BLUE};
            font-weight: 700;
        }}
        @bottom-center {{
            content: "Page " counter(page) " of " counter(pages);
            font-size: 9px;
            color: {_GRAY_500};
        }}
        @bottom-right {{
            content: "Generated {{generated_ts}}";
            font-size: 9px;
            color: {_GRAY_500};
        }}
    }}
"""


# ============================================================================
# Helper utilities
# ============================================================================

def _safe(val: Any) -> str:
    """Return a safe, HTML-escaped string representation of a value."""
    if val is None:
        return ""
    return html_escape(str(val))


def _now_eastern() -> datetime:
    """Current time in US/Eastern."""
    return datetime.now(EASTERN)


def _format_ts(dt: Optional[datetime] = None) -> str:
    """Format a datetime for display."""
    if dt is None:
        dt = _now_eastern()
    return dt.strftime("%Y-%m-%d %H:%M:%S %Z")


def _flatten_rows(data: Dict) -> List[Dict]:
    """
    Extract a flat list-of-dicts from various data shapes.

    Tries these keys in order: rows, daily_log, incidents.
    Falls back to creating a single-row dict from the top-level stats.
    """
    for key in ("rows", "daily_log", "incidents"):
        candidate = data.get(key)
        if candidate and isinstance(candidate, list):
            # Normalise list-of-lists into list-of-dicts
            if candidate and isinstance(candidate[0], (list, tuple)):
                headers = data.get("headers", [f"col_{i}" for i in range(len(candidate[0]))])
                return [dict(zip(headers, row)) for row in candidate]
            if candidate and isinstance(candidate[0], dict):
                return candidate
    # Fallback: wrap stats or metadata as a single "row"
    stats = data.get("stats")
    if stats and isinstance(stats, dict):
        return [stats]
    return []


def _dict_keys_union(rows: List[Dict]) -> List[str]:
    """Return an ordered list of all keys present across all row dicts."""
    seen = {}
    for row in rows:
        for k in row.keys():
            if k not in seen:
                seen[k] = True
    return list(seen.keys())


def _humanize_header(key: str) -> str:
    """Turn a snake_case key into a Title Case header."""
    return key.replace("_", " ").title()


# ============================================================================
# ReportRenderer
# ============================================================================

class ReportRenderer:
    """Renders report data into multiple formats (PDF, CSV, XLSX, HTML, TXT)."""

    # ------------------------------------------------------------------
    # Public entry-point
    # ------------------------------------------------------------------

    def render_all(
        self,
        run_id: int,
        template_key: str,
        title: str,
        data: Dict,
        formats: List[str],
        filters: Optional[Dict] = None,
    ) -> Dict[str, str]:
        """
        Render a report to every requested format.

        Parameters
        ----------
        run_id : int
            The report_runs.id value; used to create the artifact directory.
        template_key : str
            Determines the HTML layout (blotter, incident_summary, etc.).
        title : str
            Human-readable title shown in headers.
        data : dict
            Report payload with keys like rows, stats, metadata, incidents,
            daily_log.
        formats : list[str]
            Subset of ["pdf", "csv", "xlsx", "html", "txt"].
        filters : dict, optional
            Active filter parameters to display in the report header.

        Returns
        -------
        dict[str, str]
            Mapping of format name to the absolute path of the rendered file.
        """
        if filters is None:
            filters = {}

        artifact_dir = ensure_artifact_dir(run_id)
        generated_ts = _format_ts()
        results: Dict[str, str] = {}

        # We always need the HTML string (for html + pdf), so generate it once
        html_string: Optional[str] = None
        needs_html = any(f in formats for f in ("html", "pdf"))
        if needs_html:
            try:
                html_string = self.render_html(template_key, title, data, filters)
            except Exception:
                logger.error("HTML render failed:\n%s", traceback.format_exc())
                html_string = self._fallback_html(title, data, filters)

        # --- HTML ---
        if "html" in formats and html_string is not None:
            try:
                out = artifact_dir / f"{template_key}_{run_id}.html"
                out.write_text(html_string, encoding="utf-8")
                results["html"] = str(out.resolve())
                logger.info("HTML artifact saved: %s", out)
            except Exception:
                logger.error("Failed to write HTML artifact:\n%s", traceback.format_exc())

        # --- PDF ---
        if "pdf" in formats and html_string is not None:
            out = artifact_dir / f"{template_key}_{run_id}.pdf"
            ok = self.render_pdf(html_string, out)
            if ok:
                results["pdf"] = str(out.resolve())

        # --- CSV ---
        if "csv" in formats:
            out = artifact_dir / f"{template_key}_{run_id}.csv"
            ok = self.render_csv(data, out)
            if ok:
                results["csv"] = str(out.resolve())

        # --- XLSX ---
        if "xlsx" in formats:
            out = artifact_dir / f"{template_key}_{run_id}.xlsx"
            ok = self.render_xlsx(data, out, title=title)
            if ok:
                results["xlsx"] = str(out.resolve())

        # --- TXT ---
        if "txt" in formats:
            try:
                txt_string = self.render_txt(template_key, title, data, filters)
                out = artifact_dir / f"{template_key}_{run_id}.txt"
                out.write_text(txt_string, encoding="utf-8")
                results["txt"] = str(out.resolve())
                logger.info("TXT artifact saved: %s", out)
            except Exception:
                logger.error("Failed to write TXT artifact:\n%s", traceback.format_exc())

        logger.info(
            "render_all complete for run_id=%s template=%s formats=%s -> %s artifacts",
            run_id, template_key, formats, len(results),
        )
        return results

    # ------------------------------------------------------------------
    # HTML rendering
    # ------------------------------------------------------------------

    def render_html(
        self,
        template_key: str,
        title: str,
        data: Dict,
        filters: Optional[Dict] = None,
    ) -> str:
        """
        Render a self-contained HTML report.

        The layout varies by *template_key*; unrecognised keys fall through
        to a generic data-table layout.
        """
        if filters is None:
            filters = {}

        generated_ts = _format_ts()
        css = _BASE_CSS.replace("{{generated_ts}}", _safe(generated_ts))

        # Dispatch to specialised body builders
        body_builder = {
            "blotter": self._html_body_blotter,
            "incident_summary": self._html_body_incident_summary,
            "unit_response_stats": self._html_body_stats_table,
            "calltaker_stats": self._html_body_stats_table,
            "shift_workload": self._html_body_stats_table,
            "response_compliance": self._html_body_compliance,
            "shift_handoff": self._html_body_shift_handoff,
            "incident_detail": self._html_body_incident_detail,
            "open_incidents": self._html_body_open_incidents,
            "unit_activity": self._html_body_unit_activity,
            "unit_utilization": self._html_body_unit_utilization,
            "response_time_analysis": self._html_body_response_time_analysis,
            "monthly_summary": self._html_body_monthly_summary,
            "personnel_activity": self._html_body_stats_table,
            "incident_type_breakdown": self._html_body_stats_table,
            "location_hotspot": self._html_body_stats_table,
            "false_alarm": self._html_body_stats_table,
            "mutual_aid": self._html_body_stats_table,
            "issue_tracking": self._html_body_stats_table,
            "training_log": self._html_body_stats_table,
            # Analytics templates
            "executive_summary": self._html_body_executive_summary,
            "response_performance": self._html_body_response_performance,
            "incident_analytics": self._html_body_incident_analytics,
            "unit_performance": self._html_body_unit_performance,
            "department_overview": self._html_body_department_overview,
        }.get(template_key)

        if body_builder is None:
            # custom:* or unknown -> generic table
            body_builder = self._html_body_generic

        body_html = body_builder(template_key, title, data, filters)

        # Assemble full document
        html = (
            "<!DOCTYPE html>\n"
            "<html lang=\"en\">\n"
            "<head>\n"
            "  <meta charset=\"utf-8\">\n"
            f"  <title>{_safe(title)} - Ford Fire Department</title>\n"
            f"  <style>{css}</style>\n"
            "</head>\n"
            "<body>\n"
            "<div class=\"page-wrap\">\n"
            "  <div class=\"card\">\n"
            f"    {self._html_header(title, data)}\n"
            "    <div class=\"body\">\n"
            f"      {self._html_filter_bar(filters)}\n"
            f"      {body_html}\n"
            "    </div>\n"
            f"    {self._html_footer(generated_ts)}\n"
            "  </div>\n"
            "</div>\n"
            "</body>\n"
            "</html>"
        )
        return html

    # -- HTML fragments ------------------------------------------------

    def _html_header(self, title: str, data: Dict) -> str:
        """Professional branded header with Ford logo and metadata bar."""
        meta = data.get("metadata") or {}

        # Logo
        logo_html = ""
        if _LOGO_B64:
            logo_html = f'<img class="header-logo" src="data:image/png;base64,{_LOGO_B64}" alt="Ford">'

        # Subtitle
        subtitle = ""
        if meta.get("shift"):
            subtitle += f"Shift {_safe(meta['shift'])}"
        if meta.get("date") or data.get("date"):
            d = meta.get("date") or data.get("date")
            if subtitle:
                subtitle += " &middot; "
            subtitle += _safe(str(d))
        if not subtitle:
            subtitle = _format_ts()

        header = (
            '<div class="header">\n'
            f'  <div class="header-top">{logo_html}'
            '    <div>'
            '      <div class="header-org">Ford Motor Company &bull; BlueOval SK Battery Park &bull; Fire Department</div>'
            '      <h1>Ford Fire Department</h1>'
            '    </div>'
            '  </div>\n'
            f'  <div class="subtitle">{_safe(title)}'
            f'{"  &mdash; " + subtitle if subtitle else ""}</div>\n'
            '</div>\n'
        )

        # Metadata bar
        date_range = meta.get("date_range")
        if date_range and isinstance(date_range, list) and len(date_range) == 2:
            range_str = f"{_safe(str(date_range[0]))} to {_safe(str(date_range[1]))}"
        else:
            range_str = ""

        generated = meta.get("generated_at", _format_ts())
        tz = meta.get("timezone", "")

        meta_bar = (
            '<div class="meta-bar">\n'
            f'  <span><strong>Date Range:</strong> {range_str}</span>\n'
            f'  <span><strong>Generated:</strong> {_safe(str(generated))}</span>\n'
            f'  <span><strong>Timezone:</strong> {_safe(str(tz))}</span>\n'
            '  <span><strong>Classification:</strong> INTERNAL</span>\n'
            '</div>\n'
        )

        return header + meta_bar

    def _html_filter_bar(self, filters: Dict) -> str:
        """Render a row of filter chips (or empty string if no filters)."""
        if not filters:
            return ""
        chips = []
        for k, v in filters.items():
            if v is None or v == "":
                continue
            chips.append(
                f"<div class=\"filter-chip\"><strong>{_safe(_humanize_header(k))}:</strong> "
                f"{_safe(str(v))}</div>"
            )
        if not chips:
            return ""
        return f"<div class=\"filter-bar\">{''.join(chips)}</div>"

    def _html_footer(self, generated_ts: str) -> str:
        return (
            '<div class="branded-footer">\n'
            f'  <div class="confidential">Confidential &mdash; Ford Motor Company</div>\n'
            f'  <div style="margin-top:4px;">Generated: {_safe(generated_ts)} &nbsp;|&nbsp; '
            'FORD CAD Enterprise Reporting System</div>\n'
            '</div>\n'
        )

    def _html_stat_boxes(self, stats: Dict) -> str:
        """Render a grid of stat boxes from a stats dict."""
        if not stats:
            return ""
        boxes = []
        for k, v in stats.items():
            if isinstance(v, dict):
                continue  # skip nested dicts (e.g. incidents_by_type)
            boxes.append(
                "<div class=\"stat-box\">"
                f"<div class=\"value\">{_safe(str(v))}</div>"
                f"<div class=\"label\">{_safe(_humanize_header(k))}</div>"
                "</div>"
            )
        if not boxes:
            return ""
        return f"<div class=\"stat-grid\">{''.join(boxes)}</div>"

    def _html_data_table(self, rows: List[Dict], highlight_col: Optional[str] = None) -> str:
        """Generic table from list-of-dicts."""
        if not rows:
            return "<p style=\"text-align:center;color:#6b7280;padding:20px;\">No data available.</p>"
        headers = _dict_keys_union(rows)
        ths = "".join(f"<th>{_safe(_humanize_header(h))}</th>" for h in headers)
        trs = []
        for row in rows:
            tds = []
            for h in headers:
                val = row.get(h, "")
                cell = _safe(str(val)) if val is not None else ""
                if highlight_col and h == highlight_col:
                    css_class = ""
                    sv = str(val).lower()
                    if sv in ("pass", "true", "yes", "compliant"):
                        css_class = " class=\"pass\""
                    elif sv in ("fail", "false", "no", "non-compliant"):
                        css_class = " class=\"fail\""
                    tds.append(f"<td{css_class}>{cell}</td>")
                else:
                    tds.append(f"<td>{cell}</td>")
            trs.append(f"<tr>{''.join(tds)}</tr>")
        return f"<table><thead><tr>{ths}</tr></thead><tbody>{''.join(trs)}</tbody></table>"

    # -- Specialised body builders --------------------------------------

    def _html_body_blotter(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Chronological event list (daily log / blotter)."""
        parts: List[str] = []

        # Stats summary boxes
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        # Daily log table
        log_entries = data.get("daily_log") or data.get("rows") or []
        if log_entries:
            parts.append("<h2 class=\"section-title\">Daily Log</h2>")
            # Pick columns that are most useful for blotter view
            display_cols = ["timestamp", "category", "event_type", "unit_id",
                            "details", "created_by", "issue_found"]
            # Filter to columns that actually exist
            available = set()
            for entry in log_entries:
                available.update(entry.keys())
            cols = [c for c in display_cols if c in available]
            # Add any remaining columns
            for entry in log_entries:
                for k in entry.keys():
                    if k not in cols:
                        cols.append(k)

            ths = "".join(f"<th>{_safe(_humanize_header(c))}</th>" for c in cols)
            trs = []
            for entry in log_entries:
                tds = []
                for c in cols:
                    val = entry.get(c, "")
                    cell = _safe(str(val)) if val is not None else ""
                    style = ""
                    if c == "issue_found" and str(val) == "1":
                        style = " style=\"color:#dc2626;font-weight:700;\""
                    tds.append(f"<td{style}>{cell}</td>")
                is_issue = str(entry.get("issue_found", "0")) == "1"
                row_style = " style=\"background:#fef2f2;\"" if is_issue else ""
                trs.append(f"<tr{row_style}>{''.join(tds)}</tr>")
            parts.append(
                f"<table><thead><tr>{ths}</tr></thead><tbody>{''.join(trs)}</tbody></table>"
            )
        else:
            parts.append(
                "<p style=\"text-align:center;color:#6b7280;padding:20px;\">"
                "No daily log entries for this period.</p>"
            )

        return "\n".join(parts)

    def _html_body_incident_summary(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Incident cards with full details."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        incidents = data.get("incidents") or data.get("rows") or []
        if not incidents:
            parts.append(
                "<p style=\"text-align:center;color:#6b7280;padding:20px;\">"
                "No incidents for this period.</p>"
            )
            return "\n".join(parts)

        parts.append(f"<h2 class=\"section-title\">Incidents ({len(incidents)})</h2>")

        for inc in incidents:
            is_issue = inc.get("issue_found") in (1, "1", True)
            card_cls = "incident-card issue" if is_issue else "incident-card"
            issue_badge = (" <span class=\"badge badge-red\">ISSUE</span>" if is_issue else "")

            inc_id = inc.get("incident_number") or inc.get("incident_id") or ""
            status = inc.get("status", "")
            status_cls = "badge-green" if status == "CLOSED" else (
                "badge-yellow" if status in ("ACTIVE", "OPEN", "IN_PROGRESS") else "badge-gray"
            )

            detail_rows = ""
            for field in ("type", "location", "caller_name", "narrative",
                          "priority", "final_disposition", "created", "updated", "closed_at"):
                val = inc.get(field)
                if val:
                    detail_rows += (
                        f"<tr><td style=\"width:130px;color:{_GRAY_500};\">"
                        f"{_safe(_humanize_header(field))}:</td>"
                        f"<td>{_safe(str(val))}</td></tr>"
                    )

            # Unit assignments
            units = inc.get("units") or []
            units_html = ""
            if units:
                u_ths = ""
                u_cols = ["unit_id", "dispatched", "enroute", "arrived",
                          "cleared", "disposition"]
                available_u = set()
                for u in units:
                    available_u.update(u.keys())
                u_cols = [c for c in u_cols if c in available_u]
                u_ths = "".join(f"<th>{_safe(_humanize_header(c))}</th>" for c in u_cols)
                u_trs = []
                for u in units:
                    u_tds = "".join(f"<td>{_safe(str(u.get(c, '')))}</td>" for c in u_cols)
                    u_trs.append(f"<tr>{u_tds}</tr>")
                units_html = (
                    f"<div style=\"margin-top:10px;\"><strong>Units:</strong>"
                    f"<table><thead><tr>{u_ths}</tr></thead>"
                    f"<tbody>{''.join(u_trs)}</tbody></table></div>"
                )

            parts.append(
                f"<div class=\"{card_cls}\">"
                f"  <div style=\"display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;\">"
                f"    <strong style=\"color:{_BRAND_BLUE_DARK};font-size:14px;\">"
                f"#{_safe(str(inc_id))}{issue_badge}</strong>"
                f"    <span class=\"badge {status_cls}\">{_safe(status)}</span>"
                f"  </div>"
                f"  <table style=\"font-size:12px;\">{detail_rows}</table>"
                f"  {units_html}"
                f"</div>"
            )

        return "\n".join(parts)

    def _html_body_stats_table(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Generic stats table layout used for unit_response_stats, calltaker_stats, shift_workload."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        rows = _flatten_rows(data)
        if rows:
            parts.append(f"<h2 class=\"section-title\">Details</h2>")
            parts.append(self._html_data_table(rows))
        else:
            parts.append(
                "<p style=\"text-align:center;color:#6b7280;padding:20px;\">"
                "No detail rows available.</p>"
            )
        return "\n".join(parts)

    def _html_body_compliance(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Pass/fail compliance table with colour coding."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        rows = _flatten_rows(data)
        if rows:
            # Try to detect the pass/fail column
            compliance_col = None
            for candidate in ("compliance", "result", "pass_fail", "status", "compliant"):
                if any(candidate in r for r in rows):
                    compliance_col = candidate
                    break
            parts.append(f"<h2 class=\"section-title\">Compliance Detail</h2>")
            parts.append(self._html_data_table(rows, highlight_col=compliance_col))
        else:
            parts.append(
                "<p style=\"text-align:center;color:#6b7280;padding:20px;\">"
                "No compliance data available.</p>"
            )
        return "\n".join(parts)

    # -- New specialised body builders -----------------------------------

    def _html_body_shift_handoff(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Shift handoff report: open incidents, recent activity, and notes."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        # --- Open Incidents section (card layout) ---
        open_incidents = data.get("open_incidents") or data.get("incidents") or []
        parts.append(f'<h2 class="section-title">Open Incidents ({len(open_incidents)})</h2>')
        if open_incidents:
            for inc in open_incidents:
                inc_id = inc.get("incident_number") or inc.get("incident_id") or "N/A"
                priority = str(inc.get("priority", "")).upper()
                inc_type = inc.get("type") or inc.get("incident_type") or ""
                location = inc.get("location") or ""
                status = inc.get("status") or ""
                age = inc.get("age") or ""
                units = inc.get("units") or inc.get("assigned_units") or ""
                if isinstance(units, list):
                    units = ", ".join(str(u) for u in units)

                is_high = priority in ("HIGH", "1", "EMERGENCY")
                border_color = _BRAND_RED if is_high else _GRAY_200
                bg_color = "#fef2f2" if is_high else "#fff"
                priority_badge = "badge-red" if is_high else (
                    "badge-yellow" if priority in ("MEDIUM", "2") else "badge-gray"
                )

                parts.append(
                    f'<div class="incident-card" style="border-left:4px solid {border_color};background:{bg_color};">'
                    f'  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;">'
                    f'    <strong style="color:{_BRAND_BLUE_DARK};font-size:14px;">#{_safe(str(inc_id))}</strong>'
                    f'    <span class="badge {priority_badge}">{_safe(priority or "N/A")}</span>'
                    f'  </div>'
                    f'  <table style="font-size:12px;">'
                    f'    <tr><td style="width:120px;color:{_GRAY_500};">Type:</td><td>{_safe(inc_type)}</td></tr>'
                    f'    <tr><td style="color:{_GRAY_500};">Location:</td><td>{_safe(location)}</td></tr>'
                    f'    <tr><td style="color:{_GRAY_500};">Status:</td><td>{_safe(status)}</td></tr>'
                    f'    <tr><td style="color:{_GRAY_500};">Age:</td><td>{_safe(str(age))}</td></tr>'
                    f'    <tr><td style="color:{_GRAY_500};">Units:</td><td>{_safe(str(units))}</td></tr>'
                    f'  </table>'
                    f'</div>'
                )
        else:
            parts.append(
                '<p style="text-align:center;color:#6b7280;padding:20px;">'
                'No open incidents at time of handoff.</p>'
            )

        # --- Recent Activity table ---
        recent = data.get("recent_activity") or data.get("rows") or []
        if recent:
            parts.append(f'<h2 class="section-title">Recent Activity</h2>')
            parts.append(self._html_data_table(recent))

        # --- Notes / Pending section ---
        notes = data.get("notes") or data.get("pending") or []
        if notes:
            parts.append(f'<h2 class="section-title">Notes / Pending Items</h2>')
            if isinstance(notes, list) and notes and isinstance(notes[0], dict):
                parts.append(self._html_data_table(notes))
            elif isinstance(notes, list):
                parts.append('<ul style="margin:0;padding-left:20px;">')
                for note in notes:
                    parts.append(f'<li style="margin-bottom:4px;">{_safe(str(note))}</li>')
                parts.append('</ul>')
            else:
                parts.append(f'<p>{_safe(str(notes))}</p>')

        return "\n".join(parts)

    def _html_body_incident_detail(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """NFIRS-style printable incident detail report."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        incidents = data.get("incidents") or data.get("rows") or []
        if not incidents:
            parts.append(
                '<p style="text-align:center;color:#6b7280;padding:20px;">'
                'No incident detail available.</p>'
            )
            return "\n".join(parts)

        for idx, inc in enumerate(incidents):
            inc_id = inc.get("incident_number") or inc.get("incident_id") or "N/A"
            inc_type = inc.get("type") or inc.get("incident_type") or ""
            priority = str(inc.get("priority", "")).upper()
            status = inc.get("status") or ""

            priority_badge = "badge-red" if priority in ("HIGH", "1", "EMERGENCY") else (
                "badge-yellow" if priority in ("MEDIUM", "2") else "badge-gray"
            )
            status_cls = "badge-green" if status == "CLOSED" else (
                "badge-yellow" if status in ("ACTIVE", "OPEN", "IN_PROGRESS") else "badge-gray"
            )

            if idx > 0:
                parts.append('<div style="border-top:3px solid #1e40af;margin:32px 0;"></div>')

            # --- Incident Header ---
            parts.append(
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'margin-bottom:12px;padding:12px 16px;background:{_GRAY_100};border-radius:8px;">'
                f'  <div>'
                f'    <span style="font-size:18px;font-weight:700;color:{_BRAND_BLUE_DARK};">'
                f'Incident #{_safe(str(inc_id))}</span>'
                f'    <span class="badge {priority_badge}" style="margin-left:8px;">{_safe(priority or "N/A")}</span>'
                f'  </div>'
                f'  <div>'
                f'    <span style="margin-right:6px;font-weight:600;">{_safe(inc_type)}</span>'
                f'    <span class="badge {status_cls}">{_safe(status)}</span>'
                f'  </div>'
                f'</div>'
            )

            # --- Location / Caller block ---
            location = inc.get("location") or inc.get("address") or ""
            caller_name = inc.get("caller_name") or ""
            caller_phone = inc.get("caller_phone") or ""
            cross_street = inc.get("cross_street") or ""
            district = inc.get("district") or inc.get("zone") or ""

            loc_rows = []
            if location:
                loc_rows.append(f'<tr><td style="width:140px;color:{_GRAY_500};font-weight:600;">Location:</td><td>{_safe(location)}</td></tr>')
            if cross_street:
                loc_rows.append(f'<tr><td style="color:{_GRAY_500};font-weight:600;">Cross Street:</td><td>{_safe(cross_street)}</td></tr>')
            if district:
                loc_rows.append(f'<tr><td style="color:{_GRAY_500};font-weight:600;">District/Zone:</td><td>{_safe(district)}</td></tr>')
            if caller_name:
                loc_rows.append(f'<tr><td style="color:{_GRAY_500};font-weight:600;">Caller:</td><td>{_safe(caller_name)}</td></tr>')
            if caller_phone:
                loc_rows.append(f'<tr><td style="color:{_GRAY_500};font-weight:600;">Caller Phone:</td><td>{_safe(caller_phone)}</td></tr>')

            if loc_rows:
                parts.append(f'<h2 class="section-title">Location &amp; Caller</h2>')
                parts.append(f'<table style="font-size:12px;">{"".join(loc_rows)}</table>')

            # --- Timeline ---
            timeline_fields = [
                ("created", "Created"), ("dispatched", "Dispatched"),
                ("enroute", "En Route"), ("arrived", "Arrived"),
                ("cleared", "Cleared"), ("closed_at", "Closed"),
            ]
            tl_cells = []
            for field, label in timeline_fields:
                val = inc.get(field) or ""
                if val:
                    tl_cells.append(
                        f'<div style="text-align:center;padding:8px 14px;border:1px solid {_GRAY_200};'
                        f'border-radius:6px;background:#fff;min-width:100px;">'
                        f'<div style="font-size:10px;color:{_GRAY_500};text-transform:uppercase;">{label}</div>'
                        f'<div style="font-size:12px;font-weight:600;margin-top:2px;">{_safe(str(val))}</div>'
                        f'</div>'
                    )
            if tl_cells:
                parts.append(f'<h2 class="section-title">Timeline</h2>')
                parts.append(
                    f'<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:16px;">'
                    f'{"".join(tl_cells)}</div>'
                )

            # --- Unit Assignments table ---
            units = inc.get("units") or []
            if units:
                parts.append(f'<h2 class="section-title">Unit Assignments</h2>')
                u_cols = ["unit_id", "dispatched", "enroute", "arrived", "cleared", "disposition"]
                available_u = set()
                for u in units:
                    available_u.update(u.keys())
                u_cols = [c for c in u_cols if c in available_u]
                # Include any additional unit-level columns
                for u in units:
                    for k in u.keys():
                        if k not in u_cols:
                            u_cols.append(k)
                u_ths = "".join(f'<th>{_safe(_humanize_header(c))}</th>' for c in u_cols)
                u_trs = []
                for u in units:
                    u_tds = "".join(f'<td>{_safe(str(u.get(c, "")))}</td>' for c in u_cols)
                    u_trs.append(f'<tr>{u_tds}</tr>')
                parts.append(
                    f'<table><thead><tr>{u_ths}</tr></thead>'
                    f'<tbody>{"".join(u_trs)}</tbody></table>'
                )

            # --- Narrative entries ---
            narratives = inc.get("narratives") or inc.get("narrative") or ""
            if narratives:
                parts.append(f'<h2 class="section-title">Narrative</h2>')
                if isinstance(narratives, list):
                    for entry in narratives:
                        if isinstance(entry, dict):
                            author = entry.get("author") or entry.get("created_by") or ""
                            ts = entry.get("timestamp") or entry.get("created") or ""
                            text = entry.get("text") or entry.get("narrative") or ""
                            parts.append(
                                f'<div style="border-left:3px solid {_BRAND_BLUE_LIGHT};padding:8px 12px;'
                                f'margin-bottom:8px;background:#f8fafc;border-radius:0 6px 6px 0;">'
                                f'<div style="font-size:10px;color:{_GRAY_500};margin-bottom:2px;">'
                                f'{_safe(str(author))} &mdash; {_safe(str(ts))}</div>'
                                f'<div style="font-size:12px;">{_safe(str(text))}</div>'
                                f'</div>'
                            )
                        else:
                            parts.append(f'<p style="font-size:12px;">{_safe(str(entry))}</p>')
                else:
                    parts.append(
                        f'<div style="border-left:3px solid {_BRAND_BLUE_LIGHT};padding:8px 12px;'
                        f'background:#f8fafc;border-radius:0 6px 6px 0;font-size:12px;">'
                        f'{_safe(str(narratives))}</div>'
                    )

            # --- NFIRS fields ---
            nfirs_fields = [
                ("fire_origin", "Fire Origin"), ("fire_cause", "Cause"),
                ("fire_spread", "Fire Spread"), ("injuries", "Injuries"),
                ("fatalities", "Fatalities"), ("property_loss", "Property Loss"),
                ("contents_loss", "Contents Loss"), ("aid_given_received", "Aid Given/Received"),
                ("detector_presence", "Detector Presence"), ("sprinkler_presence", "Sprinkler Presence"),
                ("structure_type", "Structure Type"), ("construction_type", "Construction Type"),
            ]
            nfirs_rows = []
            for field, label in nfirs_fields:
                val = inc.get(field)
                if val is not None and val != "":
                    nfirs_rows.append(
                        f'<tr><td style="width:160px;color:{_GRAY_500};font-weight:600;">{label}:</td>'
                        f'<td>{_safe(str(val))}</td></tr>'
                    )
            if nfirs_rows:
                parts.append(f'<h2 class="section-title">NFIRS Data</h2>')
                parts.append(f'<table style="font-size:12px;">{"".join(nfirs_rows)}</table>')

        return "\n".join(parts)

    def _html_body_open_incidents(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Table of currently open incidents with status badges and age."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        incidents = data.get("incidents") or data.get("open_incidents") or _flatten_rows(data)
        if not incidents:
            parts.append(
                '<p style="text-align:center;color:#6b7280;padding:20px;">'
                'No open incidents.</p>'
            )
            return "\n".join(parts)

        parts.append(f'<h2 class="section-title">Open Incidents ({len(incidents)})</h2>')

        # Build table with priority colour coding
        headers = _dict_keys_union(incidents)
        ths = "".join(f'<th>{_safe(_humanize_header(h))}</th>' for h in headers)
        trs = []
        for inc in incidents:
            priority = str(inc.get("priority", "")).upper()
            is_high = priority in ("HIGH", "1", "EMERGENCY")
            is_medium = priority in ("MEDIUM", "2")
            row_bg = "#fef2f2" if is_high else ("#fffbeb" if is_medium else "")
            row_style = f' style="background:{row_bg};"' if row_bg else ""

            tds = []
            for h in headers:
                val = inc.get(h, "")
                cell = _safe(str(val)) if val is not None else ""

                # Status badge rendering
                if h == "status":
                    sv = str(val).upper()
                    badge_cls = "badge-green" if sv in ("CLOSED",) else (
                        "badge-yellow" if sv in ("ACTIVE", "OPEN", "IN_PROGRESS", "DISPATCHED") else "badge-gray"
                    )
                    cell = f'<span class="badge {badge_cls}">{cell}</span>'

                # Priority badge rendering
                if h == "priority":
                    p_cls = "badge-red" if is_high else ("badge-yellow" if is_medium else "badge-gray")
                    cell = f'<span class="badge {p_cls}">{cell}</span>'

                # Age column: bold for old incidents
                if h == "age":
                    try:
                        age_minutes = float(str(val).replace("min", "").replace("m", "").strip()) if val else 0
                        if age_minutes > 60:
                            cell = f'<strong style="color:{_BRAND_RED};">{cell}</strong>'
                    except (ValueError, TypeError):
                        pass

                tds.append(f'<td>{cell}</td>')
            trs.append(f'<tr{row_style}>{"".join(tds)}</tr>')

        parts.append(
            f'<table><thead><tr>{ths}</tr></thead><tbody>{"".join(trs)}</tbody></table>'
        )
        return "\n".join(parts)

    def _html_body_unit_activity(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Per-unit activity stats table with busiest-unit highlighting."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        rows = _flatten_rows(data)
        if not rows:
            parts.append(
                '<p style="text-align:center;color:#6b7280;padding:20px;">'
                'No unit activity data available.</p>'
            )
            return "\n".join(parts)

        parts.append(f'<h2 class="section-title">Unit Activity</h2>')

        # Determine the busiest unit by call count
        count_col = None
        for candidate in ("call_count", "total_calls", "calls", "count", "incidents"):
            if any(candidate in r for r in rows):
                count_col = candidate
                break

        busiest_val = 0
        busiest_indices = set()
        if count_col:
            for i, row in enumerate(rows):
                try:
                    v = int(row.get(count_col, 0))
                    if v > busiest_val:
                        busiest_val = v
                        busiest_indices = {i}
                    elif v == busiest_val and v > 0:
                        busiest_indices.add(i)
                except (ValueError, TypeError):
                    pass

        headers = _dict_keys_union(rows)
        ths = "".join(f'<th>{_safe(_humanize_header(h))}</th>' for h in headers)
        trs = []
        for i, row in enumerate(rows):
            is_busiest = i in busiest_indices
            row_style = f' style="background:#eff6ff;font-weight:600;"' if is_busiest else ""
            tds = []
            for h in headers:
                val = row.get(h, "")
                cell = _safe(str(val)) if val is not None else ""
                if is_busiest and h == count_col:
                    cell = f'<strong style="color:{_BRAND_BLUE_DARK};">{cell}</strong>'
                tds.append(f'<td>{cell}</td>')
            trs.append(f'<tr{row_style}>{"".join(tds)}</tr>')

        parts.append(
            f'<table><thead><tr>{ths}</tr></thead><tbody>{"".join(trs)}</tbody></table>'
        )
        return "\n".join(parts)

    def _html_body_unit_utilization(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Per-unit utilization with CSS-only progress bars inside table cells."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        rows = _flatten_rows(data)
        if not rows:
            parts.append(
                '<p style="text-align:center;color:#6b7280;padding:20px;">'
                'No unit utilization data available.</p>'
            )
            return "\n".join(parts)

        parts.append(f'<h2 class="section-title">Unit Utilization</h2>')

        # Detect the utilization/percentage column
        util_col = None
        for candidate in ("utilization", "utilization_pct", "percent", "pct", "usage", "busy_pct"):
            if any(candidate in r for r in rows):
                util_col = candidate
                break

        headers = _dict_keys_union(rows)
        ths = "".join(f'<th>{_safe(_humanize_header(h))}</th>' for h in headers)
        trs = []
        for row in rows:
            tds = []
            for h in headers:
                val = row.get(h, "")
                if h == util_col and val is not None:
                    # Render as a CSS-only bar
                    try:
                        pct = float(str(val).replace("%", "").strip())
                    except (ValueError, TypeError):
                        pct = 0
                    pct = max(0.0, min(100.0, pct))
                    bar_color = _BRAND_RED if pct >= 85 else (
                        _BRAND_YELLOW if pct >= 60 else _BRAND_GREEN
                    )
                    tds.append(
                        f'<td style="padding:4px 10px;">'
                        f'  <div style="display:flex;align-items:center;gap:8px;">'
                        f'    <div style="flex:1;background:{_GRAY_200};border-radius:4px;height:18px;overflow:hidden;">'
                        f'      <div style="width:{pct:.1f}%;height:100%;background:{bar_color};'
                        f'border-radius:4px;transition:width 0.3s;"></div>'
                        f'    </div>'
                        f'    <span style="font-size:11px;font-weight:600;min-width:42px;text-align:right;">'
                        f'{pct:.1f}%</span>'
                        f'  </div>'
                        f'</td>'
                    )
                else:
                    cell = _safe(str(val)) if val is not None else ""
                    tds.append(f'<td>{cell}</td>')
            trs.append(f'<tr>{"".join(tds)}</tr>')

        parts.append(
            f'<table><thead><tr>{ths}</tr></thead><tbody>{"".join(trs)}</tbody></table>'
        )
        return "\n".join(parts)

    def _html_body_response_time_analysis(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Detailed response time table grouped by incident type with colour coding."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        rows = _flatten_rows(data)
        if not rows:
            parts.append(
                '<p style="text-align:center;color:#6b7280;padding:20px;">'
                'No response time data available.</p>'
            )
            return "\n".join(parts)

        parts.append(f'<h2 class="section-title">Response Time Analysis</h2>')

        # Time columns to colour code
        time_cols = {
            "dispatch_time", "enroute_time", "travel_time", "total_response_time",
            "response_time", "avg_dispatch", "avg_enroute", "avg_travel", "avg_total",
        }
        # Group column detection
        group_col = None
        for candidate in ("incident_type", "type", "category", "call_type"):
            if any(candidate in r for r in rows):
                group_col = candidate
                break

        # Group rows by incident type
        grouped: Dict[str, List[Dict]] = {}
        for row in rows:
            grp = str(row.get(group_col, "All")) if group_col else "All"
            grouped.setdefault(grp, []).append(row)

        headers = _dict_keys_union(rows)

        for grp_name, grp_rows in grouped.items():
            if len(grouped) > 1:
                parts.append(
                    f'<h3 style="color:{_GRAY_700};margin:18px 0 8px;font-size:14px;">'
                    f'{_safe(grp_name)}</h3>'
                )

            ths = "".join(f'<th>{_safe(_humanize_header(h))}</th>' for h in headers)
            trs = []
            for row in grp_rows:
                tds = []
                for h in headers:
                    val = row.get(h, "")
                    cell = _safe(str(val)) if val is not None else ""
                    td_style = ""
                    if h in time_cols and val is not None:
                        try:
                            minutes = float(str(val).replace("min", "").replace("m", "").strip())
                            if minutes < 5:
                                td_style = f' style="color:{_BRAND_GREEN};font-weight:600;"'
                            elif minutes <= 8:
                                td_style = f' style="color:{_BRAND_YELLOW};font-weight:600;"'
                            else:
                                td_style = f' style="color:{_BRAND_RED};font-weight:700;"'
                        except (ValueError, TypeError):
                            pass
                    tds.append(f'<td{td_style}>{cell}</td>')
                trs.append(f'<tr>{"".join(tds)}</tr>')

            # Sub-totals row for group
            if len(grp_rows) > 1:
                subtotal_tds = []
                for h in headers:
                    if h == headers[0]:
                        subtotal_tds.append(
                            f'<td style="font-weight:700;border-top:2px solid {_GRAY_700};">Average</td>'
                        )
                    elif h in time_cols:
                        # Calculate average for time columns
                        vals = []
                        for r in grp_rows:
                            try:
                                vals.append(float(str(r.get(h, "0")).replace("min", "").replace("m", "").strip()))
                            except (ValueError, TypeError):
                                pass
                        avg = sum(vals) / len(vals) if vals else 0
                        avg_color = _BRAND_GREEN if avg < 5 else (_BRAND_YELLOW if avg <= 8 else _BRAND_RED)
                        subtotal_tds.append(
                            f'<td style="font-weight:700;border-top:2px solid {_GRAY_700};color:{avg_color};">'
                            f'{avg:.1f}</td>'
                        )
                    else:
                        subtotal_tds.append(
                            f'<td style="border-top:2px solid {_GRAY_700};"></td>'
                        )
                trs.append(f'<tr>{"".join(subtotal_tds)}</tr>')

            parts.append(
                f'<table><thead><tr>{ths}</tr></thead><tbody>{"".join(trs)}</tbody></table>'
            )

        return "\n".join(parts)

    def _html_body_monthly_summary(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Month-by-month comparison table with totals and trend arrows."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        rows = _flatten_rows(data)
        if not rows:
            parts.append(
                '<p style="text-align:center;color:#6b7280;padding:20px;">'
                'No monthly summary data available.</p>'
            )
            return "\n".join(parts)

        parts.append(f'<h2 class="section-title">Monthly Summary</h2>')

        headers = _dict_keys_union(rows)
        # Identify numeric columns for totals/averages
        numeric_cols = set()
        for h in headers:
            for row in rows:
                val = row.get(h)
                if val is not None:
                    try:
                        float(str(val).replace(",", ""))
                        numeric_cols.add(h)
                    except (ValueError, TypeError):
                        pass
                    break

        # Exclude month/period identifier columns from numeric aggregation
        label_cols = set()
        for candidate in ("month", "period", "date", "year", "label"):
            if candidate in headers:
                label_cols.add(candidate)
                numeric_cols.discard(candidate)

        ths = "".join(f'<th>{_safe(_humanize_header(h))}</th>' for h in headers)
        # Add trend column header
        ths += '<th style="text-align:center;">Trend</th>'

        trs = []
        prev_total: Optional[float] = None
        for row in rows:
            tds = []
            row_total = 0.0
            for h in headers:
                val = row.get(h, "")
                cell = _safe(str(val)) if val is not None else ""
                tds.append(f'<td>{cell}</td>')
                if h in numeric_cols and val is not None:
                    try:
                        row_total += float(str(val).replace(",", ""))
                    except (ValueError, TypeError):
                        pass

            # Trend arrow comparing row total to previous row total
            if prev_total is not None and prev_total != 0:
                diff_pct = ((row_total - prev_total) / prev_total) * 100
                if diff_pct > 5:
                    trend = f'<span style="color:{_BRAND_RED};font-size:14px;font-weight:700;">&#9650; +{diff_pct:.0f}%</span>'
                elif diff_pct < -5:
                    trend = f'<span style="color:{_BRAND_GREEN};font-size:14px;font-weight:700;">&#9660; {diff_pct:.0f}%</span>'
                else:
                    trend = f'<span style="color:{_GRAY_500};font-size:12px;">&#9644; {diff_pct:+.0f}%</span>'
            else:
                trend = '<span style="color:#9ca3af;">--</span>'
            tds.append(f'<td style="text-align:center;">{trend}</td>')
            prev_total = row_total
            trs.append(f'<tr>{"".join(tds)}</tr>')

        # Totals row
        total_tds = []
        for h in headers:
            if h == headers[0]:
                total_tds.append(
                    f'<td style="font-weight:700;border-top:2px solid {_GRAY_700};">Total</td>'
                )
            elif h in numeric_cols:
                col_sum = 0.0
                for row in rows:
                    try:
                        col_sum += float(str(row.get(h, "0")).replace(",", ""))
                    except (ValueError, TypeError):
                        pass
                total_tds.append(
                    f'<td style="font-weight:700;border-top:2px solid {_GRAY_700};">'
                    f'{col_sum:,.0f}</td>'
                )
            else:
                total_tds.append(f'<td style="border-top:2px solid {_GRAY_700};"></td>')
        total_tds.append(f'<td style="border-top:2px solid {_GRAY_700};"></td>')
        trs.append(f'<tr>{"".join(total_tds)}</tr>')

        # Averages row
        avg_tds = []
        row_count = len(rows) if rows else 1
        for h in headers:
            if h == headers[0]:
                avg_tds.append(
                    f'<td style="font-weight:700;color:{_BRAND_BLUE_DARK};">Average</td>'
                )
            elif h in numeric_cols:
                col_sum = 0.0
                for row in rows:
                    try:
                        col_sum += float(str(row.get(h, "0")).replace(",", ""))
                    except (ValueError, TypeError):
                        pass
                avg_val = col_sum / row_count
                avg_tds.append(
                    f'<td style="font-weight:700;color:{_BRAND_BLUE_DARK};">'
                    f'{avg_val:,.1f}</td>'
                )
            else:
                avg_tds.append('<td></td>')
        avg_tds.append('<td></td>')
        trs.append(f'<tr>{"".join(avg_tds)}</tr>')

        parts.append(
            f'<table><thead><tr>{ths}</tr></thead><tbody>{"".join(trs)}</tbody></table>'
        )
        return "\n".join(parts)

    # -- Chart + KPI helpers for analytics templates ------------------

    def _chart_img(self, b64: str, title: str = "", narrative: str = "") -> str:
        """Wrap a base64 chart image in a chart-section div."""
        if not b64:
            return ""
        parts = ['<div class="chart-section">']
        if title:
            parts.append(f'<h3>{_safe(title)}</h3>')
        parts.append(f'<img src="data:image/png;base64,{b64}" alt="{_safe(title)}">')
        if narrative:
            parts.append(f'<div class="chart-narrative">{_safe(narrative)}</div>')
        parts.append('</div>')
        return "\n".join(parts)

    def _kpi_cards_html(self, stats: Dict, keys: Optional[List] = None,
                        comparison: Optional[Dict] = None) -> str:
        """Render a row of branded KPI cards from stats dict."""
        if not stats:
            return ""
        items = keys if keys else list(stats.keys())
        cards = []
        for k in items:
            v = stats.get(k)
            if v is None or isinstance(v, dict):
                continue
            # Format value
            if isinstance(v, float):
                display = f"{v:.1f}"
            else:
                display = str(v)

            # Units
            if "pct" in k.lower() or "compliance" in k.lower():
                display += "%"
            elif "time" in k.lower() or "response" in k.lower() or "avg" in k.lower():
                if isinstance(v, (int, float)):
                    display += " min"

            # Trend arrow
            trend_html = ""
            if comparison:
                delta_key = k.replace("total_", "").replace("avg_", "") + "_delta_pct"
                # Try common delta naming patterns
                for try_key in [delta_key, f"incident_delta_pct", f"response_delta_pct"]:
                    delta = comparison.get(try_key)
                    if delta is not None:
                        if delta > 0:
                            trend_html = f'<div class="kpi-trend up">&#9650; +{delta:.0f}%</div>'
                        elif delta < 0:
                            trend_html = f'<div class="kpi-trend down">&#9660; {delta:.0f}%</div>'
                        else:
                            trend_html = '<div class="kpi-trend flat">&#9644; 0%</div>'
                        break

            cards.append(
                f'<div class="kpi-card">'
                f'<div class="kpi-value">{_safe(display)}</div>'
                f'<div class="kpi-label">{_safe(_humanize_header(k))}</div>'
                f'{trend_html}'
                f'</div>'
            )
        if not cards:
            return ""
        return f'<div class="kpi-row">{"".join(cards)}</div>'

    # -- Analytics body builders ----------------------------------------

    def _html_body_executive_summary(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Executive Summary Dashboard: KPI cards + trend + donut + gauge."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        comparison = data.get("comparison") or {}
        chart_data = data.get("chart_data") or {}

        # KPI cards
        kpi_keys = ["total_incidents", "avg_response_time", "p90_response_time",
                     "compliance_pct", "active_units", "issues_found"]
        parts.append(self._kpi_cards_html(stats, kpi_keys, comparison))

        # Charts
        if HAS_CHARTS and _charts:
            # Daily trend line chart
            dt = chart_data.get("daily_trend", {})
            if dt.get("labels") and dt.get("values"):
                b64 = _charts.line_chart(
                    dt["labels"],
                    [{"label": "Daily Incidents", "data": dt["values"], "color": _charts.FORD_BLUE}],
                    title="Incident Trend (Daily)"
                )
                parts.append(self._chart_img(b64, "Incident Trend",
                             f"{sum(dt['values'])} total incidents over {len(dt['labels'])} days"))

            parts.append('<div class="chart-grid">')

            # Incident type donut
            tb = chart_data.get("type_breakdown", {})
            if tb.get("labels") and tb.get("values"):
                b64 = _charts.donut_chart(tb["labels"], tb["values"], "Incident Type Breakdown")
                parts.append(self._chart_img(b64, "Incident Type Breakdown"))

            # Response time gauge
            rg = chart_data.get("response_gauge", {})
            if rg.get("target"):
                b64 = _charts.gauge_chart(rg["value"], rg["target"],
                                          "Avg Response Time", " min")
                parts.append(self._chart_img(b64, "Response Time vs Target"))

            parts.append('</div>')

        # Period comparison table
        if comparison.get("prev_total_incidents") is not None:
            parts.append('<h2 class="section-title">Period Comparison</h2>')
            parts.append('<table>')
            parts.append('<thead><tr><th>Metric</th><th>Current Period</th>'
                        '<th>Previous Period</th><th>Change</th></tr></thead><tbody>')
            rows_data = [
                ("Total Incidents", stats.get("total_incidents", 0),
                 comparison.get("prev_total_incidents", 0),
                 comparison.get("incident_delta_pct", 0)),
                ("Avg Response Time", f'{stats.get("avg_response_time", 0):.1f} min',
                 f'{comparison.get("prev_avg_response", 0) or 0:.1f} min',
                 comparison.get("response_delta_pct", 0)),
            ]
            for label, current, prev, delta in rows_data:
                delta_color = _BRAND_RED if delta > 0 else (_BRAND_GREEN if delta < 0 else _GRAY_500)
                arrow = "&#9650;" if delta > 0 else ("&#9660;" if delta < 0 else "&#9644;")
                parts.append(
                    f'<tr><td>{_safe(label)}</td><td>{_safe(str(current))}</td>'
                    f'<td>{_safe(str(prev))}</td>'
                    f'<td style="color:{delta_color};font-weight:700;">'
                    f'{arrow} {delta:+.1f}%</td></tr>'
                )
            parts.append('</tbody></table>')

        return "\n".join(parts)

    def _html_body_response_performance(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Response Performance Analysis: gauge + trend + histogram + heatmap + table."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        chart_data = data.get("chart_data") or {}

        # KPI cards
        parts.append(self._kpi_cards_html(stats))

        if HAS_CHARTS and _charts:
            # Compliance gauge
            cg = chart_data.get("compliance_gauge", {})
            if cg.get("target"):
                b64 = _charts.gauge_chart(cg["value"], cg["target"],
                                          "Response Compliance", "%")
                parts.append(self._chart_img(b64, "Compliance Rate",
                             f"{cg['value']:.0f}% of responses meet the {cg['target']}% target"))

            # Daily trend (avg + p90)
            dt = chart_data.get("daily_trend", {})
            if dt.get("labels") and dt.get("avg_values"):
                datasets = [
                    {"label": "Average", "data": dt["avg_values"], "color": _charts.FORD_BLUE},
                ]
                if dt.get("p90_values"):
                    datasets.append({"label": "90th Percentile", "data": dt["p90_values"],
                                     "color": _charts.FIRE_RED})
                b64 = _charts.line_chart(dt["labels"], datasets,
                                         "Response Time Trend (minutes)")
                parts.append(self._chart_img(b64, "Response Time Trend"))

            parts.append('<div class="chart-grid">')

            # Distribution histogram
            dist = chart_data.get("distribution", {})
            if dist.get("labels") and dist.get("values"):
                colors = [_charts.GREEN, _charts.GREEN_LIGHT, _charts.AMBER,
                          _charts.FIRE_RED_LIGHT, _charts.FIRE_RED]
                b64 = _charts.bar_chart(dist["labels"], dist["values"],
                                        "Response Time Distribution", colors[:len(dist["labels"])])
                parts.append(self._chart_img(b64, "Response Time Distribution"))

            # By type
            bt = chart_data.get("by_type", {})
            if bt.get("labels") and bt.get("values"):
                b64 = _charts.bar_chart(bt["labels"], bt["values"],
                                        "Avg Response by Incident Type")
                parts.append(self._chart_img(b64, "Response by Incident Type"))

            parts.append('</div>')

            # Heatmap
            hm = chart_data.get("heatmap", {})
            if hm.get("data") and hm.get("x_labels"):
                b64 = _charts.heatmap(hm["data"], hm["x_labels"], hm["y_labels"],
                                      "Response Time by Hour & Day of Week (minutes)")
                parts.append(self._chart_img(b64, "Activity Heatmap"))

        # Per-unit table
        rows = data.get("rows") or []
        if rows:
            parts.append('<h2 class="section-title">Per-Unit Response Metrics</h2>')
            parts.append(self._html_data_table(rows))

        return "\n".join(parts)

    def _html_body_incident_analytics(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Incident Analytics: trend + donut + hourly/DOW bars + hotspots."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        chart_data = data.get("chart_data") or {}

        parts.append(self._kpi_cards_html(stats))

        if HAS_CHARTS and _charts:
            # Daily trend
            dt = chart_data.get("daily_trend", {})
            if dt.get("labels") and dt.get("values"):
                b64 = _charts.line_chart(
                    dt["labels"],
                    [{"label": "Incidents", "data": dt["values"], "color": _charts.FORD_BLUE}],
                    "Daily Incident Volume"
                )
                parts.append(self._chart_img(b64, "Incident Volume Trend"))

            parts.append('<div class="chart-grid">')

            # Type breakdown donut
            tb = chart_data.get("type_breakdown", {})
            if tb.get("labels") and tb.get("values"):
                b64 = _charts.donut_chart(tb["labels"], tb["values"], "Incident Type Breakdown")
                parts.append(self._chart_img(b64, "Type Breakdown"))

            # Status distribution
            sd = chart_data.get("status_distribution", {})
            if sd.get("labels") and sd.get("values"):
                b64 = _charts.donut_chart(sd["labels"], sd["values"], "Status Distribution")
                parts.append(self._chart_img(b64, "Status Distribution"))

            parts.append('</div>')

            parts.append('<div class="chart-grid">')

            # Hourly pattern
            hp = chart_data.get("hourly_pattern", {})
            if hp.get("labels") and hp.get("values"):
                b64 = _charts.bar_chart(hp["labels"], hp["values"], "Incidents by Hour of Day")
                parts.append(self._chart_img(b64, "Hourly Pattern"))

            # Day of week
            dw = chart_data.get("dow_pattern", {})
            if dw.get("labels") and dw.get("values"):
                b64 = _charts.bar_chart(dw["labels"], dw["values"], "Incidents by Day of Week")
                parts.append(self._chart_img(b64, "Day of Week Pattern"))

            parts.append('</div>')

            # Location hotspots
            lh = chart_data.get("location_hotspots", {})
            if lh.get("labels") and lh.get("values"):
                b64 = _charts.horizontal_bar(lh["labels"], lh["values"],
                                             "Top 10 Incident Locations", _charts.FIRE_RED)
                parts.append(self._chart_img(b64, "Location Hotspots"))

        return "\n".join(parts)

    def _html_body_unit_performance(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Unit Performance Dashboard: utilization bars + response comparison + heatmap."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        chart_data = data.get("chart_data") or {}

        parts.append(self._kpi_cards_html(stats))

        if HAS_CHARTS and _charts:
            # Utilization horizontal bar
            util = chart_data.get("utilization", {})
            if util.get("labels") and util.get("values"):
                b64 = _charts.horizontal_bar(util["labels"], util["values"],
                                             "Unit Utilization (%)", _charts.FORD_BLUE)
                parts.append(self._chart_img(b64, "Unit Utilization"))

            parts.append('<div class="chart-grid">')

            # Call volume
            cv = chart_data.get("call_volume", {})
            if cv.get("labels") and cv.get("values"):
                b64 = _charts.bar_chart(cv["labels"], cv["values"], "Calls per Unit")
                parts.append(self._chart_img(b64, "Call Volume"))

            # Response comparison
            rc = chart_data.get("response_comparison", {})
            if rc.get("labels") and rc.get("values"):
                b64 = _charts.bar_chart(rc["labels"], rc["values"],
                                        "Avg Response Time per Unit (min)")
                parts.append(self._chart_img(b64, "Response Time Comparison"))

            parts.append('</div>')

            # Activity heatmap
            hm = chart_data.get("activity_heatmap", {})
            if hm.get("data") and hm.get("x_labels"):
                b64 = _charts.heatmap(hm["data"], hm["x_labels"], hm["y_labels"],
                                      "Unit Activity by Hour of Day")
                parts.append(self._chart_img(b64, "Activity Heatmap"))

        # Per-unit table
        rows = data.get("rows") or []
        if rows:
            parts.append('<h2 class="section-title">Unit Detail</h2>')
            parts.append(self._html_data_table(rows))

        return "\n".join(parts)

    def _html_body_department_overview(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Department Overview: multi-section comprehensive report."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        sections = data.get("sections") or {}
        chart_data = data.get("chart_data") or {}
        comparison = data.get("comparison") or {}
        recommendations = data.get("recommendations") or []

        # KPI summary
        parts.append(self._kpi_cards_html(stats, None, comparison))

        # Section 1: Executive Summary
        parts.append('<h2 class="section-title">1. Executive Summary</h2>')
        exec_data = sections.get("executive", {})
        if exec_data:
            exec_cd = exec_data.get("chart_data", {})
            if HAS_CHARTS and _charts:
                dt = exec_cd.get("daily_trend", {})
                if dt.get("labels") and dt.get("values"):
                    b64 = _charts.line_chart(
                        dt["labels"],
                        [{"label": "Daily Incidents", "data": dt["values"]}],
                        "30-Day Incident Trend"
                    )
                    parts.append(self._chart_img(b64, "Incident Trend"))

                parts.append('<div class="chart-grid">')
                tb = exec_cd.get("type_breakdown", {})
                if tb.get("labels") and tb.get("values"):
                    b64 = _charts.donut_chart(tb["labels"], tb["values"], "Incident Types")
                    parts.append(self._chart_img(b64, "Type Breakdown"))
                rg = exec_cd.get("response_gauge", {})
                if rg.get("target"):
                    b64 = _charts.gauge_chart(rg["value"], rg["target"], "Avg Response", " min")
                    parts.append(self._chart_img(b64, "Response Time"))
                parts.append('</div>')

        # Section 2: Response Performance
        parts.append('<hr class="section-divider">')
        parts.append('<h2 class="section-title">2. Response Performance</h2>')
        resp_data = sections.get("response", {})
        resp_stats = resp_data.get("stats") or {}
        if resp_stats:
            parts.append(self._kpi_cards_html(resp_stats))
        if HAS_CHARTS and _charts:
            resp_cd = resp_data.get("chart_data", {})
            rt = resp_cd.get("daily_trend", {})
            if rt.get("labels") and rt.get("avg_values"):
                datasets = [{"label": "Average", "data": rt["avg_values"], "color": _charts.FORD_BLUE}]
                if rt.get("p90_values"):
                    datasets.append({"label": "90th Pct", "data": rt["p90_values"], "color": _charts.FIRE_RED})
                b64 = _charts.line_chart(rt["labels"], datasets, "Response Time Trend")
                parts.append(self._chart_img(b64, "Response Time Trend"))
            dist = resp_cd.get("distribution", {})
            if dist.get("labels") and dist.get("values"):
                b64 = _charts.bar_chart(dist["labels"], dist["values"], "Response Time Distribution")
                parts.append(self._chart_img(b64, "Distribution"))

        # Section 3: Incident Analysis
        parts.append('<hr class="section-divider">')
        parts.append('<h2 class="section-title">3. Incident Analysis</h2>')
        inc_data = sections.get("incidents", {})
        inc_stats = inc_data.get("stats") or {}
        if inc_stats:
            parts.append(self._kpi_cards_html(inc_stats))
        if HAS_CHARTS and _charts:
            inc_cd = inc_data.get("chart_data", {})
            parts.append('<div class="chart-grid">')
            hp = inc_cd.get("hourly_pattern", {})
            if hp.get("labels") and hp.get("values"):
                b64 = _charts.bar_chart(hp["labels"], hp["values"], "Hourly Pattern")
                parts.append(self._chart_img(b64, "Hourly Pattern"))
            dw = inc_cd.get("dow_pattern", {})
            if dw.get("labels") and dw.get("values"):
                b64 = _charts.bar_chart(dw["labels"], dw["values"], "Day of Week")
                parts.append(self._chart_img(b64, "Day of Week"))
            parts.append('</div>')
            lh = inc_cd.get("location_hotspots", {})
            if lh.get("labels") and lh.get("values"):
                b64 = _charts.horizontal_bar(lh["labels"], lh["values"], "Top Locations", _charts.FIRE_RED)
                parts.append(self._chart_img(b64, "Location Hotspots"))

        # Section 4: Unit Performance
        parts.append('<hr class="section-divider">')
        parts.append('<h2 class="section-title">4. Unit Performance</h2>')
        unit_data = sections.get("units", {})
        unit_stats = unit_data.get("stats") or {}
        if unit_stats:
            parts.append(self._kpi_cards_html(unit_stats))
        if HAS_CHARTS and _charts:
            unit_cd = unit_data.get("chart_data", {})
            util_c = unit_cd.get("utilization", {})
            if util_c.get("labels") and util_c.get("values"):
                b64 = _charts.horizontal_bar(util_c["labels"], util_c["values"], "Utilization %")
                parts.append(self._chart_img(b64, "Unit Utilization"))

        unit_rows = unit_data.get("rows") or []
        if unit_rows:
            parts.append(self._html_data_table(unit_rows))

        # Section 5: Personnel Activity
        parts.append('<hr class="section-divider">')
        parts.append('<h2 class="section-title">5. Personnel Activity</h2>')
        personnel_data = sections.get("personnel", {})
        personnel_rows = personnel_data.get("rows") or []
        if personnel_rows:
            parts.append(self._html_data_table(personnel_rows[:20]))
        else:
            parts.append('<p style="color:#6b7280;text-align:center;padding:16px;">No personnel data available.</p>')

        # Section 6: Issues & Compliance
        parts.append('<hr class="section-divider">')
        parts.append('<h2 class="section-title">6. Issues &amp; Compliance</h2>')
        issues_data = sections.get("issues", {})
        issues_rows = issues_data.get("rows") or []
        if issues_rows:
            parts.append(self._html_data_table(issues_rows[:20]))
        else:
            parts.append('<p style="color:#16a34a;text-align:center;padding:16px;font-weight:600;">No issues flagged during this period.</p>')

        # Section 7: Recommendations
        parts.append('<hr class="section-divider">')
        parts.append('<h2 class="section-title">7. Recommendations</h2>')
        for rec in recommendations:
            parts.append(f'<div class="recommendation-box">{_safe(rec)}</div>')

        return "\n".join(parts)

    def _html_body_generic(self, template_key: str, title: str, data: Dict, filters: Dict) -> str:
        """Fallback generic data table for custom:* templates."""
        parts: List[str] = []
        stats = data.get("stats") or {}
        parts.append(self._html_stat_boxes(stats))

        rows = _flatten_rows(data)
        if rows:
            parts.append(self._html_data_table(rows))
        else:
            # Show raw metadata when no rows
            meta = data.get("metadata") or {}
            if meta:
                parts.append("<h2 class=\"section-title\">Metadata</h2>")
                meta_rows = [{"Key": k, "Value": str(v)} for k, v in meta.items()]
                parts.append(self._html_data_table(meta_rows))
            else:
                parts.append(
                    "<p style=\"text-align:center;color:#6b7280;padding:20px;\">"
                    "No data available for this report.</p>"
                )
        return "\n".join(parts)

    def _fallback_html(self, title: str, data: Dict, filters: Dict) -> str:
        """Ultra-minimal fallback HTML when the main renderer errors."""
        generated_ts = _format_ts()
        rows = _flatten_rows(data)
        row_count = len(rows)
        return (
            "<!DOCTYPE html><html><head><meta charset=\"utf-8\">"
            f"<title>{_safe(title)}</title></head>"
            f"<body style=\"font-family:Arial,sans-serif;padding:20px;\">"
            f"<h1>Ford Fire Department</h1>"
            f"<h2>{_safe(title)}</h2>"
            f"<p>Rows: {row_count}</p>"
            f"<p>Generated: {_safe(generated_ts)}</p>"
            f"</body></html>"
        )

    # ------------------------------------------------------------------
    # PDF rendering
    # ------------------------------------------------------------------

    def render_pdf(self, html: str, output_path: Path) -> bool:
        """
        Convert an HTML string to PDF via WeasyPrint.

        Returns True on success, False on failure (logged).
        """
        try:
            from weasyprint import HTML as WeasyprintHTML  # type: ignore
        except ImportError:
            logger.error(
                "WeasyPrint is not installed. Install with: pip install weasyprint  "
                "PDF rendering skipped for %s", output_path,
            )
            return False

        try:
            WeasyprintHTML(string=html).write_pdf(str(output_path))
            logger.info("PDF artifact saved: %s", output_path)
            return True
        except Exception:
            logger.error("PDF render failed for %s:\n%s", output_path, traceback.format_exc())
            return False

    # ------------------------------------------------------------------
    # CSV rendering
    # ------------------------------------------------------------------

    def render_csv(self, data: Dict, output_path: Path) -> bool:
        """
        Write report data rows to a CSV file.

        Handles list-of-dicts and list-of-lists (with optional ``headers``
        key in *data*).
        """
        try:
            rows = _flatten_rows(data)
            if not rows:
                logger.warning("CSV render: no rows to write for %s", output_path)
                # Write an empty CSV with a note
                output_path.write_text("No data available\n", encoding="utf-8")
                return True

            headers = _dict_keys_union(rows)

            with open(str(output_path), "w", newline="", encoding="utf-8") as fp:
                writer = csv.DictWriter(fp, fieldnames=headers, extrasaction="ignore")
                writer.writeheader()
                for row in rows:
                    # Ensure every value is a simple type
                    clean = {}
                    for h in headers:
                        v = row.get(h, "")
                        if isinstance(v, (dict, list)):
                            v = str(v)
                        clean[h] = v
                    writer.writerow(clean)

            logger.info("CSV artifact saved: %s (%d rows)", output_path, len(rows))
            return True

        except Exception:
            logger.error("CSV render failed for %s:\n%s", output_path, traceback.format_exc())
            return False

    # ------------------------------------------------------------------
    # XLSX rendering
    # ------------------------------------------------------------------

    def render_xlsx(self, data: Dict, output_path: Path, title: str = "Report") -> bool:
        """
        Render data to an XLSX workbook using openpyxl.

        Features:
        - Title row
        - Filter summary rows
        - Styled header row (bold, blue background)
        - Auto-width columns
        """
        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
        except ImportError:
            logger.error(
                "openpyxl is not installed. Install with: pip install openpyxl  "
                "XLSX rendering skipped for %s", output_path,
            )
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = title[:31]  # Excel limits sheet name to 31 chars

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            title_font = Font(bold=True, size=14, color="1E40AF")

            current_row = 1

            # -- Title --
            ws.cell(row=current_row, column=1, value=f"Ford Fire Department - {title}")
            ws.cell(row=current_row, column=1).font = title_font
            current_row += 1

            # -- Generated timestamp --
            ws.cell(row=current_row, column=1, value=f"Generated: {_format_ts()}")
            ws.cell(row=current_row, column=1).font = Font(italic=True, color="6B7280")
            current_row += 1

            # -- Filter summary --
            filters = data.get("metadata") or {}
            if filters:
                for k, v in filters.items():
                    if v is None or v == "":
                        continue
                    ws.cell(row=current_row, column=1, value=_humanize_header(str(k)))
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    ws.cell(row=current_row, column=2, value=str(v))
                    current_row += 1

            current_row += 1  # blank row separator

            # -- Data rows --
            rows = _flatten_rows(data)
            if rows:
                headers = _dict_keys_union(rows)

                # Header row
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=_humanize_header(h))
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                current_row += 1

                # Data rows
                for row_data in rows:
                    for col_idx, h in enumerate(headers, 1):
                        val = row_data.get(h, "")
                        if isinstance(val, (dict, list)):
                            val = str(val)
                        ws.cell(row=current_row, column=col_idx, value=val)
                    current_row += 1

                # Auto-width columns
                for col_idx, h in enumerate(headers, 1):
                    max_len = len(_humanize_header(h))
                    for row_data in rows[:200]:  # sample first 200 rows
                        val = str(row_data.get(h, ""))
                        if len(val) > max_len:
                            max_len = len(val)
                    ws.column_dimensions[
                        ws.cell(row=1, column=col_idx).column_letter
                    ].width = min(max_len + 4, 60)
            else:
                ws.cell(row=current_row, column=1, value="No data available")
                current_row += 1

            # -- Stats sheet (if stats present) --
            stats = data.get("stats")
            if stats and isinstance(stats, dict):
                ws_stats = wb.create_sheet(title="Summary Stats")
                ws_stats.cell(row=1, column=1, value="Metric").font = header_font
                ws_stats.cell(row=1, column=1).fill = header_fill
                ws_stats.cell(row=1, column=2, value="Value").font = header_font
                ws_stats.cell(row=1, column=2).fill = header_fill
                r = 2
                for k, v in stats.items():
                    if isinstance(v, dict):
                        # Expand nested dicts (e.g. incidents_by_type)
                        for sub_k, sub_v in v.items():
                            ws_stats.cell(row=r, column=1, value=f"{_humanize_header(k)}: {sub_k}")
                            ws_stats.cell(row=r, column=2, value=sub_v)
                            r += 1
                    else:
                        ws_stats.cell(row=r, column=1, value=_humanize_header(k))
                        ws_stats.cell(row=r, column=2, value=v)
                        r += 1
                ws_stats.column_dimensions["A"].width = 35
                ws_stats.column_dimensions["B"].width = 20

            wb.save(str(output_path))
            logger.info("XLSX artifact saved: %s (%d rows)", output_path, len(rows))
            return True

        except Exception:
            logger.error("XLSX render failed for %s:\n%s", output_path, traceback.format_exc())
            return False

    # ------------------------------------------------------------------
    # TXT rendering (for SMS / Signal / console)
    # ------------------------------------------------------------------

    def render_txt(
        self,
        template_key: str,
        title: str,
        data: Dict,
        filters: Optional[Dict] = None,
    ) -> str:
        """
        Render a concise plain-text summary suitable for SMS, Signal,
        or terminal output.

        Returns the text string (caller writes to file if needed).
        """
        if filters is None:
            filters = {}

        lines: List[str] = []
        sep = "=" * 60

        lines.append(sep)
        lines.append(f"FORD CAD - {title}")
        lines.append(sep)

        # Date range / filters
        meta = data.get("metadata") or {}
        date_range = filters.get("date_range") or meta.get("date") or data.get("date") or ""
        if date_range:
            lines.append(f"Date: {date_range}")
        shift = filters.get("shift") or meta.get("shift") or data.get("shift")
        if shift:
            lines.append(f"Shift: {shift}")

        for k, v in filters.items():
            if k in ("date_range", "shift") or v is None or v == "":
                continue
            lines.append(f"{_humanize_header(k)}: {v}")

        lines.append("")

        # Stats summary
        stats = data.get("stats") or {}
        if stats:
            lines.append("-" * 40)
            lines.append("SUMMARY")
            lines.append("-" * 40)
            for k, v in stats.items():
                if isinstance(v, dict):
                    lines.append(f"  {_humanize_header(k)}:")
                    for sub_k, sub_v in v.items():
                        lines.append(f"    {sub_k}: {sub_v}")
                else:
                    label = _humanize_header(k)
                    lines.append(f"  {label:.<30} {v}")
            lines.append("")

        # Top rows (template-dependent count)
        max_rows_map = {
            "blotter": 20,
            "incident_summary": 10,
            "unit_response_stats": 15,
            "calltaker_stats": 15,
            "shift_workload": 10,
            "response_compliance": 20,
        }
        max_rows = max_rows_map.get(template_key, 10)

        rows = _flatten_rows(data)
        if rows:
            lines.append("-" * 40)

            if template_key == "blotter":
                lines.append(f"DAILY LOG (showing {min(len(rows), max_rows)} of {len(rows)})")
                lines.append("-" * 40)
                for entry in rows[:max_rows]:
                    ts = entry.get("timestamp", "")
                    cat = entry.get("category") or entry.get("event_type", "")
                    unit = entry.get("unit_id", "")
                    details = entry.get("details", "")
                    issue = " [ISSUE]" if str(entry.get("issue_found", "0")) == "1" else ""
                    lines.append(f"  {ts}  {cat:<12} {unit:<8} {details[:60]}{issue}")

            elif template_key == "incident_summary":
                lines.append(f"INCIDENTS (showing {min(len(rows), max_rows)} of {len(rows)})")
                lines.append("-" * 40)
                for inc in rows[:max_rows]:
                    inc_id = inc.get("incident_number") or inc.get("incident_id", "?")
                    itype = inc.get("type", "")
                    loc = inc.get("location", "")
                    status = inc.get("status", "")
                    lines.append(f"  #{inc_id}  {itype}  {loc}  [{status}]")

            else:
                lines.append(f"DATA (showing {min(len(rows), max_rows)} of {len(rows)})")
                lines.append("-" * 40)
                headers = _dict_keys_union(rows)
                # Print column headers
                header_line = "  ".join(h[:12].ljust(12) for h in headers[:6])
                lines.append(f"  {header_line}")
                lines.append(f"  {'- ' * 30}")
                for row in rows[:max_rows]:
                    vals = "  ".join(str(row.get(h, ""))[:12].ljust(12) for h in headers[:6])
                    lines.append(f"  {vals}")

            if len(rows) > max_rows:
                lines.append(f"  ... and {len(rows) - max_rows} more rows")
            lines.append("")

        # Totals
        total_rows = len(rows)
        lines.append(f"Total records: {total_rows}")
        lines.append("")
        lines.append(f"Generated: {_format_ts()}")
        lines.append("Full report: [download link]")
        lines.append(sep)

        return "\n".join(lines)
